#!/usr/bin/env python3
"""
PermitPilot D2.2: Utility service territory ingestion.
Commun-ET, LLC · Internal · July 2026

Produces the data behind the intake dropdown / automatic utility sorting:
  1. utilities_by_state.json   - state -> ranked candidate utilities (dropdown source)
  2. territories_<ST>.geojson  - polygon layer per state (Tier 1 point-in-polygon)
  3. county_utility.json       - county -> utilities (Tier 3 fallback), from EIA-861

Run anywhere with internet access (not the Claude sandbox):
  pip install requests openpyxl
  python permitpilot_territory_ingest.py

Sources (verified 2026-07-15):
  PRIMARY: EIA US Energy Atlas "Electric Retail Service Territories"
           https://atlas.eia.gov -> dataset page exposes the ArcGIS REST endpoint
           (GeoServices API link on the dataset page). Paste it into ARCGIS_LAYER_URL.
  MIRROR:  Archived HIFLD Open snapshot (Sept 2024 vintage), confirmed live:
           https://maps.nccs.nasa.gov/mapping/rest/services/hifld_open/energy/FeatureServer/26
  COUNTY:  EIA Form 861 detailed files -> Service_Territory_YYYY.xlsx
           https://www.eia.gov/electricity/data/eia861/
NOTE: prefer the EIA Energy Atlas endpoint (current); the NASA mirror is the
      archived HIFLD snapshot and should be treated as backup with vintage logged.
"""

import json
import time
import urllib.parse
import urllib.request

# ---- CONFIG -----------------------------------------------------------------
ARCGIS_LAYER_URL = (
    "https://maps.nccs.nasa.gov/mapping/rest/services/hifld_open/energy/FeatureServer/26"
)  # replace with the EIA Energy Atlas endpoint for the current layer

FOOTPRINT_STATES = ["DC", "MD", "VA", "WV", "DE", "PA", "NJ", "NY", "CT", "RI",
                    "MA", "VT", "NH", "ME", "NC", "SC", "GA", "FL", "OH", "AL", "MS"]

# CET's thirteen operating relationships: pinned to the top of every dropdown.
CET_PINNED = [
    "POTOMAC ELECTRIC POWER CO",          # PEPCO
    "BALTIMORE GAS & ELECTRIC CO",        # BGE
    "CONSOLIDATED EDISON CO-NY INC",      # Con Edison
    "PUBLIC SERVICE ELEC & GAS CO",       # PSE&G
    "NIAGARA MOHAWK POWER CORP",          # National Grid upstate NY
    "MASSACHUSETTS ELECTRIC CO",          # National Grid MA
    "DUKE ENERGY CAROLINAS, LLC",
    "DUKE ENERGY PROGRESS - (NC)",
    "DUKE ENERGY FLORIDA, LLC",
    "DUKE ENERGY OHIO INC",
    "FLORIDA POWER & LIGHT CO",
    "GEORGIA POWER CO",
    "VIRGINIA ELECTRIC & POWER CO",       # Dominion Energy Virginia
    "OHIO POWER CO",                      # AEP Ohio
    "APPALACHIAN POWER CO",
    "WHEELING POWER CO",
    "ALABAMA POWER CO",
    "MISSISSIPPI POWER CO",
    "ENTERGY MISSISSIPPI LLC",
]
# Utility legal names in the federal layers differ from brand names; the list
# above uses EIA-style legal names. Verify against the pulled data and adjust:
# the matcher below is case-insensitive substring based, so partial names work.


def query_state(state: str, out_geometry: bool):
    """Query the ArcGIS layer for one state."""
    params = {
        "where": f"STATE = '{state}'",
        "outFields": "NAME,STATE,TYPE,HOLDING_CO,CUSTOMERS,REGULATED,SOURCEDATE",
        "returnGeometry": "true" if out_geometry else "false",
        "outSR": "4326",
        "f": "geojson" if out_geometry else "json",
    }
    url = ARCGIS_LAYER_URL + "/query?" + urllib.parse.urlencode(params)
    with urllib.request.urlopen(url, timeout=120) as r:
        return json.loads(r.read().decode())


def is_pinned(name: str) -> bool:
    n = (name or "").upper()
    return any(p.split(",")[0].split(" - ")[0] in n or n in p for p in CET_PINNED) or any(
        key in n for key in ["POTOMAC ELECTRIC", "BALTIMORE GAS", "CONSOLIDATED EDISON",
                             "PUBLIC SERVICE ELEC", "NIAGARA MOHAWK", "MASSACHUSETTS ELECTRIC",
                             "DUKE ENERGY", "FLORIDA POWER & LIGHT", "GEORGIA POWER",
                             "VIRGINIA ELECTRIC", "OHIO POWER", "APPALACHIAN POWER",
                             "WHEELING POWER", "ALABAMA POWER", "MISSISSIPPI POWER",
                             "ENTERGY MISSISSIPPI"])


def main():
    dropdown = {}
    for st in FOOTPRINT_STATES:
        data = query_state(st, out_geometry=False)
        feats = data.get("features", [])
        rows = []
        for f in feats:
            a = f.get("attributes", f.get("properties", {}))
            rows.append({
                "name": a.get("NAME") or a.get("name"),
                "type": a.get("TYPE") or a.get("type"),            # IOU / MUNICIPAL / COOPERATIVE
                "holding_co": a.get("HOLDING_CO") or a.get("holding_co"),
                "customers": a.get("CUSTOMERS") or a.get("customers") or 0,
                "cet_relationship": False,
            })
        for r in rows:
            r["cet_relationship"] = is_pinned(r["name"] or "")
        # Sort: CET-pinned first, then by customer count descending.
        rows.sort(key=lambda r: (not r["cet_relationship"], -(r["customers"] or 0)))
        dropdown[st] = rows
        print(f"{st}: {len(rows)} utilities ({sum(1 for r in rows if r['cet_relationship'])} pinned)")
        time.sleep(1)

        # Geometry pull for Tier 1 point-in-polygon (larger; one file per state).
        geo = query_state(st, out_geometry=True)
        with open(f"territories_{st}.geojson", "w") as fh:
            json.dump(geo, fh)

    with open("utilities_by_state.json", "w") as fh:
        json.dump({
            "generated": time.strftime("%Y-%m-%d"),
            "source": ARCGIS_LAYER_URL,
            "note": "Dropdown candidate lists. CET relationships pinned first, then by customer count. Selection must remain coordinator-confirmable per D2.2.",
            "states": dropdown,
        }, fh, indent=2)
    print("wrote utilities_by_state.json + per-state geojson")


# ---- EIA-861 county fallback (Tier 3) ---------------------------------------
def build_county_map(service_territory_xlsx: str):
    """Convert the EIA-861 Service_Territory_YYYY.xlsx into county_utility.json.
    Download the xlsx (inside the annual zip) from
    https://www.eia.gov/electricity/data/eia861/ first."""
    from openpyxl import load_workbook
    wb = load_workbook(service_territory_xlsx, read_only=True)
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:12]]
    # Expected columns include: data year, utility number, utility name, state, county
    idx = {name: header.index(name) for name in ("utility name", "state", "county") if name in header}
    counties = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            util, st, cty = row[idx["utility name"]], row[idx["state"]], row[idx["county"]]
        except Exception:
            continue
        if not (util and st and cty):
            continue
        if st not in FOOTPRINT_STATES:
            continue
        key = f"{st}:{str(cty).strip()}"
        counties.setdefault(key, [])
        if util not in counties[key]:
            counties[key].append(util)
    out = {
        k: {"utilities": sorted(v, key=lambda n: not is_pinned(n)),
            "multi_utility": len(v) > 1}
        for k, v in counties.items()
    }
    with open("county_utility.json", "w") as fh:
        json.dump(out, fh, indent=2)
    print(f"wrote county_utility.json ({len(out)} counties, "
          f"{sum(1 for v in out.values() if v['multi_utility'])} multi-utility)")


if __name__ == "__main__":
    main()
    # After downloading the EIA-861 Service Territory file, also run:
    # build_county_map("Service_Territory_2024.xlsx")
